"""Excel template processor for generating reports.

Template markers format: {{field_name}}
Supported markers are the same as DOCX processor.
"""

import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Any

import openpyxl
from openpyxl import load_workbook

from ..excel.models import StockDataBundle
from ..lstm import calculate_sma, calculate_rsi, calculate_macd, calculate_bollinger_bands


class ExcelTemplateProcessor:
    """Process Excel templates with stock data."""

    # Regex pattern to find template markers
    MARKER_PATTERN = re.compile(r"\{\{(\w+)\}\}")

    def __init__(self):
        """Initialize template processor."""
        pass

    def process(
        self,
        template_path: str,
        bundle: StockDataBundle,
        output_path: str,
        stock_name: Optional[str] = None,
    ) -> None:
        """Process template and generate report.

        Args:
            template_path: Path to the Excel template file
            bundle: Stock data bundle
            output_path: Path to save the generated report
            stock_name: Optional stock name override
        """
        wb = load_workbook(template_path)
        context = self._build_context(bundle, stock_name)

        # Process all worksheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self._process_sheet(sheet, context)

        # Save the workbook
        wb.save(output_path)

    def _build_context(
        self, bundle: StockDataBundle, stock_name: Optional[str] = None
    ) -> dict[str, Any]:
        """Build context dictionary from stock data."""
        closes = [d.close for d in bundle.data]
        volumes = [d.volume for d in bundle.data]
        turnover_rates = [d.turnover_rate for d in bundle.data]

        latest = bundle.data[-1]

        # Calculate technical indicators
        sma_5_list = calculate_sma(closes, 5)
        sma_10_list = calculate_sma(closes, 10)
        sma_20_list = calculate_sma(closes, 20)

        rsi_list = calculate_rsi(closes, 14)
        macd_dif_list, macd_dea_list, macd_hist_list = calculate_macd(closes)
        bb_upper_list, bb_middle_list, bb_lower_list = calculate_bollinger_bands(closes, 20)

        def safe_get(lst, idx, default=0.0):
            if 0 <= idx < len(lst):
                val = lst[idx]
                return val if val is not None else default
            return default

        latest_idx = len(closes) - 1

        return {
            # Basic info
            "date": datetime.now().strftime("%Y年%m月%d日"),
            "today": date.today().strftime("%Y-%m-%d"),
            "stock_code": bundle.stock_code,
            "stock_name": stock_name or bundle.stock_name or bundle.stock_code,
            # Latest price data
            "open": f"{latest.open:.2f}",
            "high": f"{latest.high:.2f}",
            "low": f"{latest.low:.2f}",
            "close": f"{latest.close:.2f}",
            "volume": f"{latest.volume:,}",
            "amount": f"{latest.amount:,.2f}",
            "turnover_rate": f"{latest.turnover_rate:.2f}",
            "change_pct": f"{latest.change_pct:.2f}",
            "change_amount": f"{latest.change_amount:.2f}",
            # Technical indicators
            "sma_5": f"{safe_get(sma_5_list, latest_idx):.2f}",
            "sma_10": f"{safe_get(sma_10_list, latest_idx):.2f}",
            "sma_20": f"{safe_get(sma_20_list, latest_idx):.2f}",
            "rsi": f"{safe_get(rsi_list, latest_idx):.2f}",
            "macd": f"{safe_get(macd_dif_list, latest_idx):.4f}",
            "macd_signal": f"{safe_get(macd_dea_list, latest_idx):.4f}",
            "macd_hist": f"{safe_get(macd_hist_list, latest_idx):.4f}",
            "bb_upper": f"{safe_get(bb_upper_list, latest_idx):.2f}",
            "bb_middle": f"{safe_get(bb_middle_list, latest_idx):.2f}",
            "bb_lower": f"{safe_get(bb_lower_list, latest_idx):.2f}",
        }

    def _process_sheet(self, sheet, context: dict) -> None:
        """Process a worksheet, replacing markers in all cells.

        Args:
            sheet: Worksheet to process
            context: Context dictionary with values
        """
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_value = self._replace_markers(cell.value, context)
                    cell.value = new_value

    def _replace_markers(self, text: str, context: dict) -> str:
        """Replace all markers in text with values from context.

        Args:
            text: Text containing markers
            context: Context dictionary

        Returns:
            Text with markers replaced
        """
        def replace_match(match):
            marker_name = match.group(1)
            value = context.get(marker_name, match.group(0))
            return str(value)

        return self.MARKER_PATTERN.sub(replace_match, text)
