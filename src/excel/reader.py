"""Excel reader for stock report generator."""

from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.cell.cell import Cell

from .models import DailyData, StockDataBundle
from .schema import COLUMNS, REQUIRED_COLUMNS, SHEET_NAME


class ValidationError(Exception):
    """Validation error with details."""

    def __init__(self, message: str, errors: Optional[list[str]] = None):
        super().__init__(message)
        self.errors = errors or []


class ValidationResult:
    """Result of validation check."""

    def __init__(self, valid: bool, errors: Optional[list[str]] = None):
        self.valid = valid
        self.errors = errors or []

    def __bool__(self) -> bool:
        return self.valid


class ExcelReader:
    """Reader for stock data Excel files.

    Format: Single stock per file with one sheet named "日线数据".
    """

    def __init__(self, file_path: str):
        """Initialize reader with file path.

        Args:
            file_path: Path to the Excel file
        """
        self.file_path = Path(file_path)
        self._workbook: Optional[openpyxl.Workbook] = None
        self._column_map: dict[str, int] = {}

    def _open(self) -> None:
        """Open the Excel file."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")
        self._workbook = openpyxl.load_workbook(self.file_path, data_only=True)

    def _close(self) -> None:
        """Close the Excel file."""
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None

    def _find_sheet(self) -> openpyxl.worksheet.worksheet.Worksheet:
        """Find and validate the required sheet.

        Returns:
            The worksheet

        Raises:
            ValidationError: If sheet is not found
        """
        if self._workbook is None:
            raise RuntimeError("Workbook not opened")

        if SHEET_NAME not in self._workbook.sheetnames:
            raise ValidationError(
                f"Sheet '{SHEET_NAME}' not found. "
                f"Available sheets: {', '.join(self._workbook.sheetnames)}"
            )

        return self._workbook[SHEET_NAME]

    def _build_column_map(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Build column index map from header row.

        Args:
            sheet: Worksheet to read from
        """
        self._column_map = {}
        headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        if headers is None:
            raise ValidationError("Header row is empty")

        for col_idx, header in enumerate(headers):
            if header is not None:
                # Find which column key this header corresponds to
                for key, config in COLUMNS.items():
                    if config["name"] == header:
                        self._column_map[key] = col_idx
                        break

    def _get_cell_value(
        self, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, col_key: str
    ) -> Optional[Cell]:
        """Get cell value by column key.

        Args:
            sheet: Worksheet
            row: Row number (1-based)
            col_key: Column key from schema

        Returns:
            Cell value or None if column not found
        """
        col_idx = self._column_map.get(col_key)
        if col_idx is None:
            return None
        return sheet.cell(row=row, column=col_idx + 1).value

    def validate(self) -> ValidationResult:
        """Validate the Excel file format.

        Returns:
            ValidationResult with valid=True if format is correct
        """
        errors: list[str] = []

        try:
            self._open()
            sheet = self._find_sheet()
            self._build_column_map(sheet)

            # Check required columns exist
            for col_key in REQUIRED_COLUMNS:
                if col_key not in self._column_map:
                    col_name = COLUMNS[col_key]["name"]
                    errors.append(f"Required column '{col_name}' not found")

            # Check data rows exist
            if sheet.max_row < 2:
                errors.append("No data rows found (header + data required)")

        except ValidationError as e:
            errors.append(str(e))
        except Exception as e:
            errors.append(f"Unexpected error: {e}")
        finally:
            self._close()

        return ValidationResult(valid=len(errors) == 0, errors=errors)

    def read(self) -> StockDataBundle:
        """Read the Excel file and return stock data bundle.

        Returns:
            StockDataBundle containing daily data

        Raises:
            ValidationError: If file format is invalid
        """
        try:
            self._open()
            sheet = self._find_sheet()
            self._build_column_map(sheet)

            # Extract stock code from first data row
            stock_code_cell = self._get_cell_value(sheet, 2, "stock_code")
            if stock_code_cell is None:
                raise ValidationError("Stock code not found in first data row")
            stock_code = str(stock_code_cell).strip()

            # Read all data rows
            daily_data: list[DailyData] = []
            for row_num in range(2, sheet.max_row + 1):
                row_data = self._read_row(sheet, row_num, stock_code)
                if row_data is not None:
                    daily_data.append(row_data)

            if not daily_data:
                raise ValidationError("No valid data rows found")

            # Sort by date
            daily_data.sort(key=lambda x: x.trade_date)

            return StockDataBundle(stock_code=stock_code, data=daily_data)

        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(f"Failed to read file: {e}")
        finally:
            self._close()

    def _read_row(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row_num: int,
        default_stock_code: str,
    ) -> Optional[DailyData]:
        """Read a single data row.

        Args:
            sheet: Worksheet
            row_num: Row number (1-based)
            default_stock_code: Stock code to use if not found in row

        Returns:
            DailyData instance or None if row is invalid
        """
        # Get trade date (required)
        trade_date_cell = self._get_cell_value(sheet, row_num, "trade_date")
        if trade_date_cell is None:
            return None

        # Parse date
        trade_date = self._parse_date(trade_date_cell)
        if trade_date is None:
            return None

        # Get stock code
        stock_code_cell = self._get_cell_value(sheet, row_num, "stock_code")
        stock_code = (
            str(stock_code_cell).strip() if stock_code_cell is not None else default_stock_code
        )

        # Get numeric values with defaults
        close = self._parse_float(self._get_cell_value(sheet, row_num, "close"))
        volume = self._parse_int(self._get_cell_value(sheet, row_num, "volume"))

        if close is None or volume is None:
            return None

        return DailyData(
            stock_code=stock_code,
            trade_date=trade_date,
            open=self._parse_float(self._get_cell_value(sheet, row_num, "open")) or 0.0,
            high=self._parse_float(self._get_cell_value(sheet, row_num, "high")) or 0.0,
            low=self._parse_float(self._get_cell_value(sheet, row_num, "low")) or 0.0,
            close=close,
            volume=volume,
            amount=self._parse_float(self._get_cell_value(sheet, row_num, "amount")) or 0.0,
            turnover_rate=self._parse_float(self._get_cell_value(sheet, row_num, "turnover_rate"))
            or 0.0,
            change_pct=self._parse_float(self._get_cell_value(sheet, row_num, "change_pct"))
            or 0.0,
            change_amount=self._parse_float(
                self._get_cell_value(sheet, row_num, "change_amount")
            )
            or 0.0,
        )

    def _parse_date(self, value) -> Optional[date]:
        """Parse date from cell value."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def _parse_float(self, value) -> Optional[float]:
        """Parse float from cell value."""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    def _parse_int(self, value) -> Optional[int]:
        """Parse int from cell value."""
        if value is None:
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return None
